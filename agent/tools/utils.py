"""Utility tools for the agent."""
import json
import os
from datetime import datetime

from langchain_core.tools import tool


@tool
def get_current_time() -> str:
    """Get the current date and time."""
    now = datetime.now()
    return now.strftime("%Y-%m-%d %H:%M:%S %A")


@tool
def write_excel(filename: str, headers: str, rows: str) -> str:
    """Write data to an Excel .xlsx file.

    Args:
        filename: Output filename, e.g. 'hot_search.xlsx'. Saved under workspace.
        headers: JSON array of column header strings, e.g. '["排名","标题","热度"]'
        rows: JSON array of rows, each row a JSON array, e.g. '[["1","xxx","123万"],["2","yyy","98万"]]'
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    header_list = json.loads(headers)
    row_list = json.loads(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Style header
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(header_list, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data rows
    for row_idx, row_data in enumerate(row_list, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # Auto-fit column widths
    for col_idx in range(1, len(header_list) + 1):
        max_width = 0
        for row_idx in range(1, len(row_list) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_width = max(max_width, len(str(cell_value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_width + 4, 60)

    # Save to workspace root
    save_path = os.path.join(os.getcwd(), filename)
    wb.save(save_path)
    return f"Excel file saved: {save_path}\nHeaders: {header_list}\nRow count: {len(row_list)}"


def get_utility_tools():
    """Return all utility tools."""
    return [get_current_time, write_excel]
